from django.shortcuts import render
from django.http import HttpResponse
import pandas as pd
import openpyxl
import requests
import re
from io import BytesIO

def check_domain(domain):
    domain = domain.strip()

    headers = {"User-Agent": "Mozilla/5.0"}

    # Prepare both versions
    if not domain.startswith(("http://", "https://")):
        https_url = "https://" + domain
        http_url = "http://" + domain
    else:
        https_url = domain
        http_url = domain

    try:
        # Try HTTPS first
        response = requests.get(
            https_url,
            timeout=5,
            headers=headers,
            allow_redirects=False
        )
        ssl_status = "Yes"
        final_domain = https_url

    except requests.exceptions.RequestException:
        try:
            # Fallback to HTTP
            response = requests.get(
                http_url,
                timeout=5,
                headers=headers,
                allow_redirects=False
            )
            ssl_status = "No"
            final_domain = http_url

        except:
            return {
                "domain": domain,
                "status": "Connection Error",
                "ssl": "No"
            }

    code = response.status_code

    # Status
    if 200 <= code < 300:
        status = "Success"
    elif 300 <= code < 400:
        status = "Redirect"
    elif 400 <= code < 500:
        status = "Client Error"
    elif 500 <= code < 600:
        status = "Server Error"
    else:
        status = "Unknown"

    # Description
    STATUS_DESCRIPTIONS = {
        200: "Success",
        201: "Created",
        204: "No Content",
        301: "Moved Permanently",
        302: "Found (Redirect)",
        304: "Not Modified",
        400: "Bad Request",
        401: "Unauthorized",
        403: "Forbidden",
        404: "Not Found",
        405: "Method Not Allowed",
        408: "Request Timeout",
        429: "Too Many Requests",
        500: "Internal Server Error",
        502: "Bad Gateway",
        503: "Service Unavailable",
        504: "Gateway Timeout"
    }

    description = STATUS_DESCRIPTIONS.get(code, "Other Response")

    return {
        "domain": final_domain,
        "status": status,
        "code": code,
        "description": description,
        "time": round(response.elapsed.total_seconds(), 2),
        "ssl": ssl_status
    }


# def check_domain(domain):
#     domain = domain.strip()
#
#     if not domain.startswith(("http://", "https://")):
#         domain = "https://" + domain
#
#     try:
#         headers = {"User-Agent": "Mozilla/5.0"}
#
#         response = requests.get(domain, timeout=5, headers=headers, allow_redirects=False)
#
#         code = response.status_code
#
#         # Status
#         if 200 <= code < 300:
#             status = "Success"
#         elif 300 <= code < 400:
#             status = "Redirect"
#         elif 400 <= code < 500:
#             status = "Client Error"
#         elif 500 <= code < 600:
#             status = "Server Error"
#         else:
#             status = "Unknown"
#
#         ssl_status = "Yes" if domain.startswith("https://") else "No"
#
#         # Description
#         STATUS_DESCRIPTIONS = {
#             200: "Success",
#             201: "Created",
#             204: "No Content",
#             301: "Moved Permanently",
#             302: "Found (Redirect)",
#             304: "Not Modified",
#             400: "Bad Request",
#             401: "Unauthorized",
#             403: "Forbidden",
#             404: "Not Found",
#             405: "Method Not Allowed",
#             408: "Request Timeout",
#             429: "Too Many Requests",
#             500: "Internal Server Error",
#             502: "Bad Gateway",
#             503: "Service Unavailable",
#             504: "Gateway Timeout"
#         }
#
#         description = STATUS_DESCRIPTIONS.get(code, "Other Response")
#
#         return {
#             "domain": domain,
#             "status": status,
#             "code": code,
#             "description": description,
#             "time": round(response.elapsed.total_seconds(), 2),
#             "ssl": ssl_status,
#         }
#
#     except requests.exceptions.SSLError:
#         return {"domain": domain, "status": status, "ssl": "No"}
#
#     except requests.exceptions.RequestException:
#         return {"domain": domain, "status": "Connection Error", "ssl": "No"}


def is_valid_domain(domain):
    pattern = r'^(https?:\/\/)?([a-zA-Z0-9-]+\.)+[a-zA-Z]{2,}'
    # return re.match(pattern, domain)
    return bool(re.match(pattern, domain))

def home(request):
    results = []
    domain_input = ""
    file_name = ""

    if request.method == "POST":
        domains = []

        domain_input = request.POST.get("domains", "")

        if domain_input:
            domains.extend(domain_input.splitlines())

        uploaded_file = request.FILES.get("file")

        if uploaded_file:

            if not uploaded_file.name.endswith('.xlsx'):
                return render(request, "home.html", {
                    "error": "Only Excel (.xlsx) files allowed"
                })

            file_name = uploaded_file.name

            try:
                df = pd.read_excel(uploaded_file)

                file_domains = []

                possible_cols = []

                for col in df.columns:
                    col_name = str(col).lower()

                    if any(keyword in col_name for keyword in ["domain", "url", "site", "website"]):
                        possible_cols.append(col)

                if possible_cols:
                    for col in possible_cols:
                        for value in df[col].dropna():
                            value = str(value).strip()

                            if is_valid_domain(value):
                                file_domains.append(value)

                else:
                    for col in df.columns:
                        for value in df[col].dropna():
                            value = str(value).strip()

                            if is_valid_domain(value):
                                file_domains.append(value)

                file_domains = list(set(file_domains))

                domains.extend(file_domains)

            except Exception as e:
                print("File error:", e)

        if len(domains) > 100:
            return render(request, "home.html", {
                "error": "Maximum 100 domains allowed"
            })

        for d in domains:

            d = str(d).strip()

            if "127.0.0.1" in d or "localhost" in d:
                continue

            if d and is_valid_domain(d):
                results.append(check_domain(d))

        request.session["results"] = results

    if request.method == "GET":
        request.session.pop("results", None)
        results = []

    return render(request, "home.html", {
        "results": results,
        "domain_input": domain_input,
        "file_name": file_name
    })


# def download_excel(request):
#     results = request.session.get("results", [])
#
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Report"
#
#     ws.append(["Domain", "Status", "Code", "Response Time"])
#
#     for r in results:
#         ws.append([
#             r.get("domain"),
#             r.get("status"),
#             r.get("code", "-"),
#             r.get("time", "-")
#         ])
#
#     ws.auto_filter.ref = ws.dimensions
#
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename=report.xlsx'
#
#     wb.save(response)
#     return response

def download_excel(request):
    results = request.session.get("results", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.append(["Domain", "Status", "Code", "Response Time", "SSL"])

    for r in results:
        ws.append([
            r.get("domain"),
            r.get("status"),
            r.get("code", "-"),
            r.get("time", "-"),
            r.get("ssl", "-")
        ])

    ws.auto_filter.ref = f"A1:E{ws.max_row}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=report.xlsx'

    response.write(output.read())
    return response